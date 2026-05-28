"""Combine the 'Calibration' and 'Pressure Gauge' sheets of the source workbook
into a single CSV with a unified schema.

Run from repo root:
    python scripts/xlsx_to_csv.py

Inputs  : data/source.xlsx
Outputs : data.csv (next to index.html so the dashboard can fetch it)

Output schema:
    description       Free text - instrument or machine name
    equipment_type    Upper-case category (CALIPER, PRESSURE, ...)
    vendor            Cal-by party (Sendi Mahir, Trescal, internal, ...)
    cost              Number, RM, 2 decimals
    last_cal_date     ISO date YYYY-MM-DD (empty if missing)
    area              Plant area / department
    status            Item Status / Item Type from the source
    source            'Calibration' | 'Pressure Gauge' - lets users filter
"""

from __future__ import annotations

import csv
import sys
from datetime import datetime, date
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "data" / "source.xlsx"
OUT = ROOT / "data.csv"

CALIBRATION_MAP = {
    "Equipment Description": "description",
    "Equipment Type": "equipment_type",
    "Cal by": "vendor",
    "Cal. Cost": "cost",
    "Last Cal Date": "last_cal_date",
    "Area": "area",
    "Item Status": "status",
}

PRESSURE_MAP = {
    "Machine No": "description",
    "Equipment Type": "equipment_type",
    "Cal. By": "vendor",
    "Cal Cost (RM)": "cost",
    "Last Cal Date": "last_cal_date",
    "Area": "area",
    "Item Type": "status",
}

OUTPUT_FIELDS = [
    "description",
    "equipment_type",
    "vendor",
    "cost",
    "last_cal_date",
    "area",
    "status",
    "source",
]


def to_iso_date(v) -> str:
    if v is None or v == "":
        return ""
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    try:
        return datetime.fromisoformat(str(v)).date().isoformat()
    except (ValueError, TypeError):
        return ""


def clean_str(v) -> str:
    return "" if v is None else str(v).strip()


def clean_cost(v) -> float:
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return round(float(v), 2)
    s = str(v).upper().replace("RM", "").replace(",", "").strip()
    try:
        return round(float(s), 2)
    except ValueError:
        return 0.0


def read_sheet(wb: openpyxl.Workbook, sheet_name: str, mapping: dict, source_label: str):
    if sheet_name not in wb.sheetnames:
        print(f"[warn] sheet '{sheet_name}' not in workbook - skipped", file=sys.stderr)
        return

    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    headers = [clean_str(h) for h in next(rows)]

    col_idx = {}
    for src_header, dest_field in mapping.items():
        try:
            col_idx[dest_field] = headers.index(src_header)
        except ValueError:
            print(
                f"[warn] sheet '{sheet_name}' missing expected column '{src_header}'",
                file=sys.stderr,
            )

    kept = 0
    skipped = 0
    for raw in rows:
        if all(v is None or v == "" for v in raw):
            skipped += 1
            continue

        row = {f: "" for f in OUTPUT_FIELDS}
        row["source"] = source_label

        for dest_field, idx in col_idx.items():
            val = raw[idx] if idx < len(raw) else None
            if dest_field == "cost":
                row[dest_field] = clean_cost(val)
            elif dest_field == "last_cal_date":
                row[dest_field] = to_iso_date(val)
            elif dest_field == "equipment_type":
                row[dest_field] = clean_str(val).upper()
            else:
                row[dest_field] = clean_str(val)

        if not row["description"] and not row["equipment_type"]:
            skipped += 1
            continue

        if not isinstance(row["cost"], (int, float)):
            row["cost"] = 0.0

        kept += 1
        yield row

    print(f"[info] {sheet_name}: kept={kept} skipped={skipped}", file=sys.stderr)


def main() -> int:
    if not SRC.exists():
        print(f"[error] source not found: {SRC}", file=sys.stderr)
        return 1

    OUT.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(SRC, data_only=True, read_only=True)

    all_rows = []
    all_rows.extend(read_sheet(wb, "Calibration", CALIBRATION_MAP, "Calibration"))
    all_rows.extend(read_sheet(wb, "Pressure Gauge", PRESSURE_MAP, "Pressure Gauge"))

    # Stable sort for clean git diffs
    all_rows.sort(key=lambda r: (r["last_cal_date"] or "", r["description"]))

    with OUT.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=OUTPUT_FIELDS)
        writer.writeheader()
        writer.writerows(all_rows)

    total_cost = sum(r["cost"] for r in all_rows)
    with_dates = sum(1 for r in all_rows if r["last_cal_date"])

    print(f"[ok] wrote {OUT.relative_to(ROOT)}", file=sys.stderr)
    print(f"     rows         : {len(all_rows)}", file=sys.stderr)
    print(f"     rows w/ date : {with_dates}", file=sys.stderr)
    print(f"     total cost   : RM {total_cost:,.2f}", file=sys.stderr)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
